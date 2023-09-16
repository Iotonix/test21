import re
from django.core.exceptions import MultipleObjectsReturned
from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework import views
from rest_framework.parsers import FileUploadParser, MultiPartParser, FormParser
from rest_framework.decorators import api_view, permission_classes
from api.serializers import (
    LanguageSerializer,
    PlatformSerializer,
    BaseSerializer,
    TranslationSerializer,
)
from api.models.language import Language
from api.models.platform import Platform
from api.models.incoming import TextRequest
from api.models.base import Base
from api.models.uniquetext import UniqueText
from rest_framework.permissions import IsAuthenticated, AllowAny
from api.models.base_text_set import BaseText
from api.models.translation import Translation
import pandas as pd
from pandas import DataFrame as df


class FileUploadView(views.APIView):
    # parser_classes = (FileUploadParser,)
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [
        IsAuthenticated,
    ]

    def put(self, request, format=None):
        file_obj = request.data["filename"]
        # file_data = file_obj.read()
        # book = xlrd.open_workbook(file_contents=file_data)
        wb = load_workbook(file_obj, data_only=True)
        resp_data = []
        print(wb.sheetnames)
        resp_data.append({"Sheets": wb.sheetnames})
        print("-----------------")

        for sheet in wb.worksheets:
            print(sheet.title)
            try:
                platform = Platform.objects.get(name=sheet.title)
                print("Platform:", platform.name)
                code = self.parseBase(sheet, platform)
                if code < 400:
                    code = self.parseLanguageSheets(sheet, platform)
            except Platform.DoesNotExist:
                platform = None
                code = 400

        return Response(status=code)

    def parseLanguageSheets(self, sheet, platform):
        """TODOC"""
        start_col = platform.translation_start_col
        print("parsing languages from: ", start_col)
        code = 201
        for column in sheet.iter_cols(
            min_row=2, max_row=2, min_col=start_col, max_col=sheet.max_column
        ):
            for cell in column:
                if cell.value is None:
                    print("checking Platform: ", platform.name)
                    print("Cell Value is None! :-( Continue")
                    continue
                try:
                    language = None
                    print("checking Platform: ", platform.name)
                    print("Language: ", cell.value.strip())
                    lingo = cell.value.strip()
                    language = Language.objects.get(locale__iexact=lingo)
                except Language.DoesNotExist:
                    code = 400
                    print("language Not Found")
                if language is not None:
                    print("We found the language", language.name_en)
                    code = self.handleLanguage(cell, sheet, platform, language)
        return code

    def handleLanguage(self, cell, sheet, platform, language):
        """TODOC"""
        cCol = cell.column
        text_col = platform.default_col
        # for c in sheet[cCol]:
        # for c in sheet.iter_rows(min_col=cCol, max_col=cCol, min_row=platform.data_start_row):
        for c_lingo, c_default in zip(
            sheet.iter_rows(
                min_col=cCol, max_col=cCol, min_row=platform.data_start_row
            ),
            sheet.iter_rows(
                min_col=text_col, max_col=text_col, min_row=platform.data_start_row
            ),
        ):
            language_text = c_lingo[0].value
            default_text = c_default[0].value
            print(language_text, default_text)
            if language_text is None:
                continue
            try:
                text = UniqueText.objects.get(textlabel=default_text)
            except UniqueText.DoesNotExist:
                print("Unique Text Not Found: ", default_text)
                return 404

            if language_text is not None:
                translation, ceated = Translation.objects.get_or_create(
                    language=language, uniquetext=text, trans=language_text
                )
                if ceated:
                    print(
                        "Created Translation. language="
                        + language.name_en
                        + " unique="
                        + default_text
                        + " val=",
                        language_text,
                    )
                else:
                    print(
                        "Fetched Translation. language="
                        + language.name_en
                        + " unique="
                        + default_text
                        + " val=",
                        language_text,
                    )
                    pass
        print("Finished:", language)
        return 201

    def parseBase(self, sheet, platform):
        """TODOC"""
        print("ParsingBase: ", platform.name)
        start_row = platform.data_start_row
        print("Start Row is: ", start_row)
        key_col = platform.key_col
        print("Key Column is: ", key_col)
        default_col = platform.default_col
        print("Default Column is: ", default_col)
        for row in sheet.iter_rows(
            min_row=start_row, max_row=sheet.max_row, min_col=key_col, max_col=key_col
        ):
            for cell in row:
                if cell.value is None:
                    break
                default_text = sheet.cell(row=cell.row, column=default_col).value
                if default_text is None:
                    print(
                        "Row: ",
                        cell.row,
                        " Coll",
                        default_col,
                        " has no value for key: ",
                        cell.value,
                    )
                    break

                key = cell.value
                # print(cell.value, "=", default_text)
                base, created = Base.objects.get_or_create(platform=platform, key=key)
                msg = " created" if created else " fetched"
                # print("BASE: ", base.key, msg)

                text, created = UniqueText.objects.get_or_create(textlabel=default_text)
                msg = " created" if created else " fetched"
                # print("TEXT: ",text.textlabel, msg)

                base_text_set, created = BaseText.objects.get_or_create(
                    base=base, uniquetext=text
                )
                msg = " created" if created else " fetched"
                if created:
                    print("BASETEXT: ", text.textlabel, msg)
        print("Done setting the Base!")
        return 201


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def langUpload(request):
    fileRequest = request.data.get("file")
    platform = request.data.get("platform")
    fileObject = df(pd.read_csv(fileRequest))
    try:
        pf = Platform.objects.get(name=platform)
    except:
        return Response("Platform not in list!", status=400)
    languages = fileObject.columns
    for row in fileObject.values:
        if row[0]:
            base, create = Base.objects.get_or_create(platform=pf, key=row[0])
            text, create = UniqueText.objects.get_or_create(textlabel=row[1])
            BaseText.objects.get_or_create(base=base, uniquetext=text)
            for index, language in enumerate(languages):
                try:
                    lang = Language.objects.get(locale=language)
                except Language.DoesNotExist:
                    continue
                try:
                    trans, _ = Translation.objects.get_or_create(
                        language=lang, uniquetext=text
                    )
                    trans.trans = row[index]
                    trans.save()
                    print(text, "-", trans)
                except MultipleObjectsReturned:
                    pass
    return Response("Upload Successful", status=200)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def langAdd(request):
    lang = request.data.get("lang")
    if lang:
        qwe = Language.objects.get_or_create(locale=lang)
        return Response("Yes", status=200)
    else:
        return Response("No language given", status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def langGet(request):
    return Response([lang.locale for lang in Language.objects.all()], status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def langNew(request):
    text = request.data.get("text")
    print("adding new text:", text)
    qwe = TextRequest.objects.get_or_create(text=text)
    print("########################################")
    return Response(text, status=200)


# @api_view(['GET'])
# @permission_classes([AllowAny])
# def getIncoming(request):
#     qwe = TextRequest.objects.all()
#     incomingTexts = [text.text for text in qwe]
#     return Response(incomingTexts, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def getIncoming(request):
    language = Language.objects.get(locale=request.query_params.get("lang"))
    data = Translation.objects.filter(language=language)
    serializer = TranslationSerializer(data, many=True)
    return Response(serializer.data, status=200)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def updateTranslation(request):
    print(request.data)
    language = Language.objects.filter(locale=request.data.get("lang")).first()
    print(language)
    if language:
        text = UniqueText.objects.filter(textlabel=request.data.get("label")).first()
        print(text)
        trans = Translation.objects.filter(language=language, uniquetext=text).first()
        if trans:
            trans.trans = request.data.get("translation")
            trans.save()
            return Response("Translation updated", status=200)
        else:
            return Response("No text found!", status=400)
    else:
        return Response("No language found!", status=400)
